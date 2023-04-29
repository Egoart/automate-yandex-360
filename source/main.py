import openpyxl
import os
import pprint
import sys

from add_new_user import create_user
from departments import fetch_depatments
from create_login import generate_named_login
from validations import format_phone_number
from validations import handle_dept_input
from validations import handle_email
from validations import handle_input_options
from validations import handle_login
from validations import handle_phone_input
from validations import handle_user_inputs
from utils import get_or_create_file_path

DATA_FILE = "email_form_v1.1.xlsx"

department_dict = fetch_depatments()


def main():
    data_file_path = get_or_create_file_path(DATA_FILE)
    if os.path.exists(data_file_path):
        print("Обнаружен файл с данными")
        handle_data_file(data_file_path)
    else:
        handle_user_data()


def handle_data_file(file_path):
    wb = openpyxl.load_workbook(file_path)
    ws = wb.active
    get_dept_key_by_value = [
        key for key, value in department_dict.items() if value == ws["F4"].value
    ][0]

    raw_user_data = {
        "initiator_email": ws["B1"].value,
        "last_name": ws["A4"].value,
        "first_name": ws["B4"].value,
        "middle_name": ws["C4"].value,
        "position": ws["D4"].value,
        "phone": format_phone_number(ws["E4"].value),
    }
    user_data = {
        k: (v.strip() if isinstance(v, str) else "не указан")
        for (k, v) in raw_user_data.items()
    }
    user_data.update({"department_id": get_dept_key_by_value})
    choose_login_type_confirm_create_email(user_data)


def handle_user_data():
    initiator_email = handle_email("Введите email заявителя: ")
    user_inputs = {
        "last_name": "Введите фамилию сотрудника: ",
        "first_name": "Введите имя: ",
        "middle_name": "Введите отчество: ",
        "position": "Введите должность: ",
    }
    user_data_values = [
        handle_user_inputs(user_input).capitalize()
        for user_input in user_inputs.values()
    ]
    user_data = dict(zip(user_inputs.keys(), user_data_values))
    user_phone = handle_phone_input(
        "Введите номер телефона в формате +375 ХХ ХХХ ХХ ХХ: "
    )
    print(f"Список подразделений предприятия:\n {pprint.pformat(department_dict)}")
    department_id = handle_dept_input(
        "Укажите номер подраздления из списка выше: ", len(department_dict)
    )
    user_data.update({"initiator_email": initiator_email})
    user_data.update({"phone": user_phone})
    user_data.update({"department_id": department_id})
    choose_login_type_confirm_create_email(user_data)


def choose_login_type_confirm_create_email(user_data):
    user_login = handle_login_type(
        last_name=user_data.get("last_name").upper(),
        first_name=user_data.get("first_name").upper(),
    )
    user_data.update(user_login)
    confirm_user_data(user_data)
    create_user(user_data)


def handle_login_type(last_name: str, first_name: str) -> dict:
    login_type = handle_input_options("Cоздать произвольный login? да/нет: ")
    if login_type == "да":
        user_login = handle_login("Введите login латинскими буквами: ")
    elif login_type == "нет":
        user_login = generate_named_login(last_name, first_name)
    return {"login": user_login.lower()}


def confirm_user_data(user_data):
    print(
        f"Будет создан пользователь со следующими параметрами:\n {pprint.pformat(user_data)}"
    )
    answer_type = handle_input_options("Данные верны, продолжить? да/нет:")
    if answer_type == "да":
        print("Данные используются для создания почтового ящика...")
    elif answer_type == "нет":
        print(
            "Создание почтового ящика прервано. Перезапустите скрипт еще раз и введите корректные данные."
        )
        sys.exit()


if __name__ == "__main__":
    main()
